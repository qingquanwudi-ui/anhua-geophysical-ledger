import argparse
import json
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter


def load_workbook_meta(conn, workbook_id):
    row = conn.execute(
        "SELECT id, file_name FROM template_workbook WHERE id = ?",
        (workbook_id,),
    ).fetchone()
    if not row:
        raise ValueError(f"未找到模板入库ID：{workbook_id}")
    return row


def apply_style(cell, style_json):
    if not style_json:
        return
    style = json.loads(style_json)
    font = style.get("font", {})
    fill = style.get("fill", {})
    alignment = style.get("alignment", {})
    border = style.get("border", {})
    protection = style.get("protection", {})

    cell.number_format = style.get("number_format") or "General"
    cell.font = Font(
        name=font.get("name"),
        size=font.get("size"),
        bold=font.get("bold"),
        italic=font.get("italic"),
        underline=font.get("underline"),
        color=font.get("color"),
    )
    if fill.get("type") and fill.get("fgColor"):
        cell.fill = PatternFill(fill_type=fill.get("type"), fgColor=fill.get("fgColor"))
    elif fill.get("type"):
        cell.fill = PatternFill(fill_type=fill.get("type"))
    cell.alignment = Alignment(
        horizontal=alignment.get("horizontal"),
        vertical=alignment.get("vertical"),
        wrap_text=alignment.get("wrap_text"),
        text_rotation=alignment.get("text_rotation") or 0,
    )
    cell.border = Border(
        left=Side(style=border.get("left")),
        right=Side(style=border.get("right")),
        top=Side(style=border.get("top")),
        bottom=Side(style=border.get("bottom")),
    )
    cell.protection = Protection(
        locked=protection.get("locked", True),
        hidden=protection.get("hidden", False),
    )


def restore_template(db_path, workbook_id, output_path):
    conn = sqlite3.connect(db_path)
    try:
        load_workbook_meta(conn, workbook_id)
        wb = Workbook()
        default_ws = wb.active
        wb.remove(default_ws)

        sheets = conn.execute(
            """
            SELECT id, sheet_index, sheet_name, freeze_panes, sheet_state
            FROM template_sheet
            WHERE workbook_id = ?
            ORDER BY sheet_index
            """,
            (workbook_id,),
        ).fetchall()

        for sheet_id, _, sheet_name, freeze_panes, sheet_state in sheets:
            ws = wb.create_sheet(sheet_name)
            ws.freeze_panes = freeze_panes
            ws.sheet_state = sheet_state

            for row_index, height, hidden, outline_level in conn.execute(
                """
                SELECT row_index, height, hidden, outline_level
                FROM template_row_dimension
                WHERE sheet_id = ?
                """,
                (sheet_id,),
            ):
                dim = ws.row_dimensions[row_index]
                dim.height = height
                dim.hidden = bool(hidden)
                dim.outlineLevel = outline_level or 0

            for col_index, col_letter, width, hidden, outline_level in conn.execute(
                """
                SELECT col_index, col_letter, width, hidden, outline_level
                FROM template_column_dimension
                WHERE sheet_id = ?
                """,
                (sheet_id,),
            ):
                letter = col_letter or get_column_letter(col_index)
                dim = ws.column_dimensions[letter]
                dim.width = width
                dim.hidden = bool(hidden)
                dim.outlineLevel = outline_level or 0

            for cell_ref, raw_value, formula, style_json in conn.execute(
                """
                SELECT cell_ref, raw_value, formula, style_json
                FROM template_cell
                WHERE sheet_id = ?
                ORDER BY row_index, col_index
                """,
                (sheet_id,),
            ):
                cell = ws[cell_ref]
                cell.value = formula or raw_value
                apply_style(cell, style_json)

            for (range_ref,) in conn.execute(
                "SELECT range_ref FROM template_merge WHERE sheet_id = ? ORDER BY id",
                (sheet_id,),
            ):
                ws.merge_cells(range_ref)
    finally:
        conn.close()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="从模板数据库恢复Excel台账")
    parser.add_argument("--db", default="outputs/ledger_template.db", help="模板SQLite数据库")
    parser.add_argument("--workbook-id", type=int, default=1, help="模板入库ID")
    parser.add_argument("-o", "--output", default="outputs/恢复台账模板.xlsx", help="输出Excel路径")
    args = parser.parse_args()
    restore_template(Path(args.db), args.workbook_id, Path(args.output))
    print(f"恢复文件：{Path(args.output).resolve()}")


if __name__ == "__main__":
    main()
