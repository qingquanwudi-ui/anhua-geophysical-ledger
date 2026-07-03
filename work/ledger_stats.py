import argparse
import datetime as dt
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


COMMON_FIELD_ALIASES = {
    "委托单位": ["委托单位"],
    "工程名称": ["工程名称"],
    "单位工程": ["单位工程"],
    "分部工程": ["分部工程"],
    "单元工程": ["单元工程"],
    "工程部位": ["工程部位"],
    "委托编号": ["委托编号"],
    "报告编号": ["报告编号"],
    "注浆日期": ["注浆日期"],
    "委托日期": ["委托日期"],
    "检测日期": ["检测日期"],
    "报告日期": ["报告日期"],
    "批准日期": ["批准日期", "批准时间"],
    "检测结果": ["检测结果"],
    "施工数量": ["施工数量", "施工数量（根）", "施工数量(根）"],
    "检测数量": ["检测数量", "检测数量（根）"],
    "备注": ["备注"],
}


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value).strip()
    return value


def normalize_header(value):
    text = str(clean_text(value))
    return text.replace(" ", "").replace("　", "")


def normalize_date(value):
    if value in (None, "", "/"):
        return ""
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value

    text = str(value).strip()
    if not text or text == "/":
        return ""

    patterns = [
        r"(20\d{2})[-/](\d{1,2})[-/](\d{1,2})",
        r"(20\d{2})年(\d{1,2})月(\d{1,2})日?",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            year, month, day = map(int, match.groups())
            return dt.date(year, month, day)
    return text


def find_header_row(ws):
    for row in range(1, min(ws.max_row, 20) + 1):
        values = [normalize_header(ws.cell(row, col).value) for col in range(1, min(ws.max_column, 30) + 1)]
        if "序号" in values and "委托单位" in values and "报告编号" in values:
            return row
    return None


def read_context_value(ws, label):
    for row in range(1, min(ws.max_row, 8) + 1):
        for col in range(1, min(ws.max_column, 5) + 1):
            value = normalize_header(ws.cell(row, col).value)
            if value.startswith(label):
                return clean_text(ws.cell(row, col + 1).value)
    return ""


def build_header_map(ws, header_row):
    header_map = {}
    for col in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(header_row, col).value)
        if header and header not in header_map:
            header_map[header] = col
    return header_map


def display_header(value, col):
    text = str(clean_text(value))
    if text:
        return text
    return f"未命名列{get_column_letter(col)}"


def build_header_labels(ws, header_row):
    labels = {}
    previous = ""
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(header_row, col).value
        current = display_header(raw, col)
        if current.startswith("未命名列") and previous:
            current = f"{previous}_{get_column_letter(col)}"
        labels[col] = current
        if raw not in (None, ""):
            previous = current
    return labels


def find_col(header_map, aliases):
    normalized_aliases = [normalize_header(alias) for alias in aliases]
    for alias in normalized_aliases:
        if alias in header_map:
            return header_map[alias]
    for header, col in header_map.items():
        if any(alias in header for alias in normalized_aliases):
            return col
    return None


def is_data_row(ws, row):
    value = ws.cell(row, 1).value
    if value is None:
        return False
    text = str(value).strip()
    return bool(re.fullmatch(r"\d+(\.0)?", text))


def sheet_detection_type(sheet_name, title):
    if title:
        return str(title).replace("台账", "").replace("检测", "检测").strip()
    return re.sub(r"^\d+[_-]?", "", sheet_name).strip()


def extract_rows(input_path):
    wb = load_workbook(input_path, data_only=True)
    rows = []
    skipped = []

    for ws in wb.worksheets:
        header_row = find_header_row(ws)
        if not header_row:
            skipped.append((ws.title, "未识别到表头"))
            continue

        header_map = build_header_map(ws, header_row)
        header_labels = build_header_labels(ws, header_row)
        title = clean_text(ws.cell(1, 1).value)
        detection_type = sheet_detection_type(ws.title, title)
        detection_org = read_context_value(ws, "检测单位") or read_context_value(ws, "试验单位")
        construction_org = read_context_value(ws, "施工单位")

        field_cols = {
            field: find_col(header_map, aliases)
            for field, aliases in COMMON_FIELD_ALIASES.items()
        }

        for row in range(header_row + 1, ws.max_row + 1):
            if not is_data_row(ws, row):
                continue

            item = {
                "来源文件": Path(input_path).name,
                "工作表": ws.title,
                "检测类型": detection_type,
                "检测单位": detection_org,
                "施工单位": construction_org,
                "原始行号": row,
            }
            for field, col in field_cols.items():
                value = ws.cell(row, col).value if col else ""
                if "日期" in field:
                    value = normalize_date(value)
                else:
                    value = clean_text(value)
                item[field] = value

            used_cols = {col for col in field_cols.values() if col}
            extras = []
            for col in range(1, ws.max_column + 1):
                if col in used_cols:
                    continue
                value = clean_text(ws.cell(row, col).value)
                if value in (None, ""):
                    continue
                extras.append(
                    {
                        "字段名": header_labels.get(col, f"未命名列{get_column_letter(col)}"),
                        "字段值": value,
                        "列号": col,
                        "列名": get_column_letter(col),
                    }
                )
            item["扩展字段"] = extras

            rows.append(item)

    return rows, skipped


def month_key(value):
    if isinstance(value, dt.date):
        return f"{value.year:04d}-{value.month:02d}"
    if isinstance(value, str) and re.fullmatch(r"20\d{2}-\d{2}-\d{2}", value):
        return value[:7]
    return "未填写"


def number_value(value):
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return 0.0
    match = re.search(r"-?\d+(\.\d+)?", str(value))
    return float(match.group(0)) if match else 0.0


def build_summaries(rows):
    by_type = defaultdict(lambda: {"记录数": 0, "合格": 0, "不合格": 0, "施工数量": 0.0, "检测数量": 0.0})
    by_month = defaultdict(lambda: {"记录数": 0, "合格": 0, "不合格": 0, "检测数量": 0.0})

    for row in rows:
        result = str(row.get("检测结果", ""))
        type_key = row.get("检测类型") or "未分类"
        month = month_key(row.get("报告日期") or row.get("检测日期") or row.get("委托日期"))

        for target in (by_type[type_key], by_month[month]):
            target["记录数"] += 1
            if "不" in result:
                target["不合格"] += 1
            elif "合格" in result:
                target["合格"] += 1

        by_type[type_key]["施工数量"] += number_value(row.get("施工数量"))
        by_type[type_key]["检测数量"] += number_value(row.get("检测数量"))
        by_month[month]["检测数量"] += number_value(row.get("检测数量"))

    return by_type, by_month


def safe_sheet_name(name, used_names):
    cleaned = re.sub(r"[:\\/?*\[\]]", "_", str(name or "未分类"))
    cleaned = cleaned[:31] or "未分类"
    candidate = cleaned
    index = 1
    while candidate in used_names:
        suffix = f"_{index}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        index += 1
    used_names.add(candidate)
    return candidate


def sort_value_for_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str):
        parsed = normalize_date(value)
        if isinstance(parsed, dt.date):
            return parsed
    return dt.date.max


def sort_rows_by_detection_date(rows):
    return sorted(
        rows,
        key=lambda row: (
            sort_value_for_date(row.get("检测日期")),
            str(row.get("检测类型", "")),
            str(row.get("报告编号", "")),
            row.get("原始行号", 0),
        ),
    )


def write_sheet(ws, headers, records):
    ws.append(headers)
    for record in records:
        ws.append([record.get(header, "") for header in headers])

    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, len(headers) + 1):
        width = min(max(len(str(headers[col - 1])) + 4, 12), 36)
        ws.column_dimensions[get_column_letter(col)].width = width


def type_extra_headers(rows):
    headers = []
    seen = set()
    for row in rows:
        for extra in row.get("扩展字段", []):
            name = extra.get("字段名")
            if name and name not in seen:
                seen.add(name)
                headers.append(name)
    return headers


def rows_with_extras(rows, base_headers, extra_headers):
    output = []
    for row in rows:
        extra_map = {}
        for extra in row.get("扩展字段", []):
            name = extra.get("字段名")
            if name and name not in extra_map:
                extra_map[name] = extra.get("字段值")
        output_row = {header: row.get(header, "") for header in base_headers}
        output_row.update({header: extra_map.get(header, "") for header in extra_headers})
        output.append(output_row)
    return output


def export_workbook(rows, skipped, output_path):
    wb = Workbook()
    used_sheet_names = set()
    sorted_rows = sort_rows_by_detection_date(rows)
    detail = wb.active
    detail.title = "台账明细"
    used_sheet_names.add(detail.title)

    detail_headers = [
        "来源文件", "工作表", "检测类型", "检测单位", "施工单位", "原始行号",
        "委托单位", "工程名称", "单位工程", "分部工程", "单元工程", "工程部位",
        "委托编号", "报告编号", "注浆日期", "委托日期", "检测日期", "报告日期",
        "批准日期", "检测结果", "施工数量", "检测数量", "备注",
    ]
    write_sheet(detail, detail_headers, sorted_rows)

    by_type, by_month = build_summaries(sorted_rows)

    type_records = []
    for name, data in sorted(by_type.items()):
        record = {"检测类型": name, **data}
        record["合格率"] = data["合格"] / data["记录数"] if data["记录数"] else 0
        record["检测比例"] = data["检测数量"] / data["施工数量"] if data["施工数量"] else ""
        type_records.append(record)

    ws_type = wb.create_sheet("检测类型汇总")
    used_sheet_names.add(ws_type.title)
    write_sheet(ws_type, ["检测类型", "记录数", "合格", "不合格", "施工数量", "检测数量", "合格率", "检测比例"], type_records)
    for row in ws_type.iter_rows(min_row=2, min_col=7, max_col=8):
        for cell in row:
            cell.number_format = "0.00%"

    month_records = [{"月份": name, **data} for name, data in sorted(by_month.items())]
    ws_month = wb.create_sheet("月度汇总")
    used_sheet_names.add(ws_month.title)
    write_sheet(ws_month, ["月份", "记录数", "合格", "不合格", "检测数量"], month_records)

    item_records = []
    rows_by_type = defaultdict(list)
    for row in sorted_rows:
        rows_by_type[row.get("检测类型") or "未分类"].append(row)

    for detection_type, type_rows in sorted(rows_by_type.items()):
        type_rows = sort_rows_by_detection_date(type_rows)
        unit_summary = defaultdict(lambda: {"记录数": 0, "合格": 0, "不合格": 0, "检测数量": 0.0})
        for row in type_rows:
            key = row.get("单位工程") or "未填写"
            result = str(row.get("检测结果", ""))
            unit_summary[key]["记录数"] += 1
            unit_summary[key]["检测数量"] += number_value(row.get("检测数量"))
            if "不" in result:
                unit_summary[key]["不合格"] += 1
            elif "合格" in result:
                unit_summary[key]["合格"] += 1

        for unit_project, data in sorted(unit_summary.items()):
            item_records.append({"检测类型": detection_type, "单位工程": unit_project, **data})

        sheet_name = safe_sheet_name(f"明细_{detection_type}", used_sheet_names)
        ws_item = wb.create_sheet(sheet_name)
        extra_headers = type_extra_headers(type_rows)
        item_headers = detail_headers + extra_headers
        write_sheet(ws_item, item_headers, rows_with_extras(type_rows, detail_headers, extra_headers))

    ws_item_summary = wb.create_sheet("检测项统计")
    used_sheet_names.add(ws_item_summary.title)
    write_sheet(ws_item_summary, ["检测类型", "单位工程", "记录数", "合格", "不合格", "检测数量"], item_records)

    if skipped:
        ws_skip = wb.create_sheet("未导入工作表")
        used_sheet_names.add(ws_skip.title)
        write_sheet(ws_skip, ["工作表", "原因"], [{"工作表": name, "原因": reason} for name, reason in skipped])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="抽水蓄能检测台账基础统计工具")
    parser.add_argument("input", help="输入Excel台账文件路径")
    parser.add_argument("-o", "--output", default="outputs/台账统计结果.xlsx", help="输出Excel文件路径")
    args = parser.parse_args()

    rows, skipped = extract_rows(Path(args.input))
    export_workbook(rows, skipped, Path(args.output))
    print(f"导入明细记录：{len(rows)}")
    print(f"跳过工作表：{len(skipped)}")
    print(f"输出文件：{Path(args.output).resolve()}")


if __name__ == "__main__":
    main()
