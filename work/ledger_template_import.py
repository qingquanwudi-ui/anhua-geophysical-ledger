import argparse
import datetime as dt
import json
import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def to_text(value):
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value)


def style_to_json(cell):
    style = {
        "number_format": cell.number_format,
        "font": {
            "name": cell.font.name,
            "size": cell.font.sz,
            "bold": cell.font.bold,
            "italic": cell.font.italic,
            "underline": cell.font.underline,
            "color": cell.font.color.rgb if cell.font.color and cell.font.color.type == "rgb" else None,
        },
        "fill": {
            "type": cell.fill.fill_type,
            "fgColor": cell.fill.fgColor.rgb if cell.fill.fgColor and cell.fill.fgColor.type == "rgb" else None,
        },
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical,
            "wrap_text": cell.alignment.wrap_text,
            "text_rotation": cell.alignment.text_rotation,
        },
        "border": {
            "left": cell.border.left.style,
            "right": cell.border.right.style,
            "top": cell.border.top.style,
            "bottom": cell.border.bottom.style,
        },
        "protection": {
            "locked": cell.protection.locked,
            "hidden": cell.protection.hidden,
        },
    }
    return json.dumps(style, ensure_ascii=False)


def connect(db_path):
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db(conn):
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS template_workbook (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL,
            file_name TEXT NOT NULL,
            imported_at TEXT NOT NULL,
            project_name TEXT,
            section_name TEXT,
            source_type TEXT,
            discipline TEXT,
            remark TEXT
        );

        CREATE TABLE IF NOT EXISTS template_sheet (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            workbook_id INTEGER NOT NULL REFERENCES template_workbook(id) ON DELETE CASCADE,
            sheet_index INTEGER NOT NULL,
            sheet_name TEXT NOT NULL,
            max_row INTEGER,
            max_column INTEGER,
            freeze_panes TEXT,
            sheet_state TEXT
        );

        CREATE TABLE IF NOT EXISTS template_cell (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sheet_id INTEGER NOT NULL REFERENCES template_sheet(id) ON DELETE CASCADE,
            row_index INTEGER NOT NULL,
            col_index INTEGER NOT NULL,
            cell_ref TEXT NOT NULL,
            raw_value TEXT,
            formula TEXT,
            data_type TEXT,
            number_format TEXT,
            style_json TEXT
        );

        CREATE TABLE IF NOT EXISTS template_merge (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sheet_id INTEGER NOT NULL REFERENCES template_sheet(id) ON DELETE CASCADE,
            range_ref TEXT NOT NULL,
            min_row INTEGER,
            min_col INTEGER,
            max_row INTEGER,
            max_col INTEGER
        );

        CREATE TABLE IF NOT EXISTS template_row_dimension (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sheet_id INTEGER NOT NULL REFERENCES template_sheet(id) ON DELETE CASCADE,
            row_index INTEGER NOT NULL,
            height REAL,
            hidden INTEGER,
            outline_level INTEGER
        );

        CREATE TABLE IF NOT EXISTS template_column_dimension (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sheet_id INTEGER NOT NULL REFERENCES template_sheet(id) ON DELETE CASCADE,
            col_index INTEGER NOT NULL,
            col_letter TEXT NOT NULL,
            width REAL,
            hidden INTEGER,
            outline_level INTEGER
        );

        CREATE INDEX IF NOT EXISTS idx_template_cell_sheet_pos
            ON template_cell(sheet_id, row_index, col_index);
        CREATE INDEX IF NOT EXISTS idx_template_sheet_workbook
            ON template_sheet(workbook_id, sheet_index);
        """
    )


def insert_workbook(conn, input_path, args):
    now = dt.datetime.now().isoformat(timespec="seconds")
    cur = conn.execute(
        """
        INSERT INTO template_workbook (
            source_file, file_name, imported_at, project_name, section_name,
            source_type, discipline, remark
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            str(input_path),
            input_path.name,
            now,
            args.project,
            args.section,
            args.source_type,
            args.discipline,
            args.remark,
        ),
    )
    return cur.lastrowid


def insert_sheet(conn, workbook_id, ws, sheet_index):
    cur = conn.execute(
        """
        INSERT INTO template_sheet (
            workbook_id, sheet_index, sheet_name, max_row, max_column,
            freeze_panes, sheet_state
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            workbook_id,
            sheet_index,
            ws.title,
            ws.max_row,
            ws.max_column,
            str(ws.freeze_panes) if ws.freeze_panes else None,
            ws.sheet_state,
        ),
    )
    return cur.lastrowid


def insert_cells(conn, sheet_id, ws, keep_blank):
    count = 0
    for row in ws.iter_rows():
        for cell in row:
            has_value = cell.value is not None
            has_style = cell.style_id != 0
            if not keep_blank and not has_value:
                continue
            formula = cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None
            conn.execute(
                """
                INSERT INTO template_cell (
                    sheet_id, row_index, col_index, cell_ref, raw_value,
                    formula, data_type, number_format, style_json
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    sheet_id,
                    cell.row,
                    cell.column,
                    cell.coordinate,
                    to_text(cell.value),
                    formula,
                    cell.data_type,
                    cell.number_format,
                    style_to_json(cell) if has_style or has_value else None,
                ),
            )
            count += 1
    return count


def insert_merges(conn, sheet_id, ws):
    count = 0
    for merged_range in ws.merged_cells.ranges:
        conn.execute(
            """
            INSERT INTO template_merge (
                sheet_id, range_ref, min_row, min_col, max_row, max_col
            )
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                sheet_id,
                str(merged_range),
                merged_range.min_row,
                merged_range.min_col,
                merged_range.max_row,
                merged_range.max_col,
            ),
        )
        count += 1
    return count


def insert_dimensions(conn, sheet_id, ws):
    row_count = 0
    for row_index, dim in ws.row_dimensions.items():
        if dim.height is None and not dim.hidden and not dim.outlineLevel:
            continue
        conn.execute(
            """
            INSERT INTO template_row_dimension (
                sheet_id, row_index, height, hidden, outline_level
            )
            VALUES (?, ?, ?, ?, ?)
            """,
            (sheet_id, row_index, dim.height, int(bool(dim.hidden)), dim.outlineLevel or 0),
        )
        row_count += 1

    col_count = 0
    for col_letter, dim in ws.column_dimensions.items():
        col_index = openpyxl_col_index(col_letter)
        conn.execute(
            """
            INSERT INTO template_column_dimension (
                sheet_id, col_index, col_letter, width, hidden, outline_level
            )
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (sheet_id, col_index, col_letter, dim.width, int(bool(dim.hidden)), dim.outlineLevel or 0),
        )
        col_count += 1
    return row_count, col_count


def openpyxl_col_index(col_letter):
    index = 0
    for char in col_letter:
        if not char.isalpha():
            continue
        index = index * 26 + (ord(char.upper()) - ord("A") + 1)
    return index


def import_template(input_path, db_path, args):
    wb = load_workbook(input_path, data_only=False)
    conn = connect(db_path)
    totals = {"sheets": 0, "cells": 0, "merges": 0, "row_dims": 0, "col_dims": 0}
    try:
        init_db(conn)
        with conn:
            workbook_id = insert_workbook(conn, input_path, args)
            for index, ws in enumerate(wb.worksheets, start=1):
                sheet_id = insert_sheet(conn, workbook_id, ws, index)
                totals["sheets"] += 1
                totals["cells"] += insert_cells(conn, sheet_id, ws, args.keep_blank_cells)
                totals["merges"] += insert_merges(conn, sheet_id, ws)
                row_dims, col_dims = insert_dimensions(conn, sheet_id, ws)
                totals["row_dims"] += row_dims
                totals["col_dims"] += col_dims
        return workbook_id, totals
    finally:
        conn.close()


def main():
    parser = argparse.ArgumentParser(description="原样录入Excel台账模板信息到SQLite")
    parser.add_argument("input", help="输入Excel文件路径")
    parser.add_argument("--db", default="outputs/ledger_template.db", help="SQLite数据库路径")
    parser.add_argument("--project", default="湖南安化抽水蓄能电站", help="项目名称")
    parser.add_argument("--section", default="Q2标", help="标段名称")
    parser.add_argument("--source-type", default="施工检测", help="资料来源")
    parser.add_argument("--discipline", default="物探", help="专业")
    parser.add_argument("--remark", default="", help="备注")
    parser.add_argument(
        "--keep-blank-cells",
        action="store_true",
        help="是否保存空白单元格。默认只保存有值单元格，合并、列宽、行高仍会保存。",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    db_path = Path(args.db)
    if not input_path.exists():
        print(f"输入文件不存在：{input_path}", file=sys.stderr)
        sys.exit(1)

    workbook_id, totals = import_template(input_path, db_path, args)
    print(f"模板入库ID：{workbook_id}")
    print(f"工作表数量：{totals['sheets']}")
    print(f"单元格记录：{totals['cells']}")
    print(f"合并区域：{totals['merges']}")
    print(f"行高记录：{totals['row_dims']}")
    print(f"列宽记录：{totals['col_dims']}")
    print(f"数据库文件：{db_path.resolve()}")


if __name__ == "__main__":
    main()
