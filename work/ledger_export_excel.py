import argparse
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


EXPORT_COLUMNS = [
    ("project_name", "项目名称"),
    ("section_name", "标段名称"),
    ("source_type", "资料来源"),
    ("discipline", "检测专业"),
    ("detection_type", "检测类型"),
    ("detection_unit", "检测单位"),
    ("construction_unit", "施工单位"),
    ("unit_project", "单位工程"),
    ("sub_project", "分部工程"),
    ("item_project", "单元工程"),
    ("work_part", "工程部位"),
    ("entrust_no", "委托编号"),
    ("report_no", "报告编号"),
    ("entrust_date", "委托日期"),
    ("detection_date", "检测日期"),
    ("report_date", "报告日期"),
    ("result", "检测结果"),
    ("construction_qty", "施工数量"),
    ("detection_qty", "检测数量"),
    ("source_file", "来源文件"),
    ("source_sheet", "来源工作表"),
    ("source_row", "来源行号"),
]


def build_where(args):
    clauses = []
    params = []
    if args.section:
        clauses.append("section_name = ?")
        params.append(args.section)
    if args.source_type:
        clauses.append("source_type = ?")
        params.append(args.source_type)
    if args.detection_type:
        clauses.append("detection_type LIKE ?")
        params.append(f"%{args.detection_type}%")
    if args.date_from:
        clauses.append("detection_date >= ?")
        params.append(args.date_from)
    if args.date_to:
        clauses.append("detection_date <= ?")
        params.append(args.date_to)
    if args.result:
        clauses.append("result LIKE ?")
        params.append(f"%{args.result}%")
    return ("WHERE " + " AND ".join(clauses)) if clauses else "", params


def query_rows(conn, args):
    where_sql, params = build_where(args)
    select_sql = ", ".join(col for col, _ in EXPORT_COLUMNS)
    sql = f"""
        SELECT {select_sql}
        FROM detection_ledger
        {where_sql}
        ORDER BY
            CASE WHEN detection_date IS NULL THEN 1 ELSE 0 END,
            detection_date ASC,
            detection_type ASC,
            report_no ASC,
            source_row ASC
    """
    return conn.execute(sql, params).fetchall()


def write_excel(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "查询结果"
    headers = [title for _, title in EXPORT_COLUMNS]
    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="从SQLite台账库查询并导出Excel")
    parser.add_argument("--db", default="outputs/ledger.db", help="SQLite数据库路径")
    parser.add_argument("-o", "--output", default="outputs/台账查询结果.xlsx", help="导出Excel路径")
    parser.add_argument("--section", default="", help="标段筛选")
    parser.add_argument("--source-type", default="", help="资料来源筛选")
    parser.add_argument("--detection-type", default="", help="检测类型关键词")
    parser.add_argument("--date-from", default="", help="检测日期起始，格式 YYYY-MM-DD")
    parser.add_argument("--date-to", default="", help="检测日期截止，格式 YYYY-MM-DD")
    parser.add_argument("--result", default="", help="检测结果关键词")
    args = parser.parse_args()

    conn = sqlite3.connect(args.db)
    try:
        rows = query_rows(conn, args)
    finally:
        conn.close()
    write_excel(rows, Path(args.output))
    print(f"导出记录：{len(rows)}")
    print(f"输出文件：{Path(args.output).resolve()}")


if __name__ == "__main__":
    main()
